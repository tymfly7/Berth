"""
Model comparison reporting
==========================
Builds the styled Excel model-comparison workbook and loads per-model training
details from the training history JSON / YOLO results CSVs. Pure data/reporting
logic — kept out of the web entrypoint so main.py stays an API surface.
"""

import csv
import io
import json

import config


def build_comparison_excel(comparison: list) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model Comparison"

    # ── palette ───────────────────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill(fgColor=hex_color, fill_type="solid")

    HDR_BG   = fill("3730a3")   # indigo-700
    CNN_BG   = fill("dbeafe")   # blue-100
    YCLS_BG  = fill("fef3c7")   # amber-100
    YDET_BG  = fill("ede9fe")   # violet-100
    BEST_BG  = fill("bbf7d0")   # green-200
    NA_BG    = fill("f1f5f9")   # slate-100

    thin   = Side(style="thin",   color="94a3b8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    hdr_font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    best_font = Font(bold=True, size=10, name="Calibri")
    note_font = Font(italic=True, size=9, color="64748b", name="Calibri")

    # ── column definitions ────────────────────────────────────────────────────
    # (header text, width, data key, format)
    COLS = [
        ("Model",            20, "model",            None),
        ("Type",             14, "type",              None),
        ("Epochs",            8, "epochs",            "0"),
        ("Train Time (s)",   12, "train_time",        "0.0"),
        ("Best Val Acc (%)", 13, "best_val_acc",      "0.00"),
        ("Test Acc (%)",     12, "test_accuracy",     "0.00"),
        ("Precision (%)",    12, "test_precision",    "0.00"),
        ("Recall (%)",       12, "test_recall",       "0.00"),
        ("F1 Score (%)",     12, "test_f1",           "0.00"),
        ("Total Params",     13, "total_params",      "#,##0"),
        ("Trainable Params", 14, "trainable_params",  "#,##0"),
    ]

    MODEL_LABELS = {
        "cnn_scratch":     "CNN Scratch",
        "resnet50":        "ResNet-50",
        "mobilenetv4s":     "MobileNetV4",
        "yolo26_classify": "YOLO26 Classify",
        "yolo26":          "YOLO26 Detect",
    }
    ROW_BG = {
        "cnn_scratch":     CNN_BG,
        "resnet50":        CNN_BG,
        "mobilenetv4s":     CNN_BG,
        "yolo26_classify": YCLS_BG,
        "yolo26":          YDET_BG,
    }

    from openpyxl.utils import get_column_letter

    # ── header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    for ci, (hdr, width, _, _) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill      = HDR_BG
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── data rows ─────────────────────────────────────────────────────────────
    for ri, r in enumerate(comparison, 2):
        name   = r.get("model", "")
        row_bg = ROW_BG.get(name, fill("FFFFFF"))
        ws.row_dimensions[ri].height = 20

        for ci, (_, _, key, fmt) in enumerate(COLS, 1):
            if key == "model":
                val = MODEL_LABELS.get(name, name)
                align = left
            elif key == "type":
                val   = str(r.get("type", "classification")).title()
                align = center
            else:
                raw = r.get(key)
                val = round(raw, 2) if isinstance(raw, float) else raw
                align = center

            cell = ws.cell(row=ri, column=ci, value=val if val is not None else None)
            cell.font      = body_font
            cell.alignment = align
            cell.border    = border

            if val is None:
                cell.value = "—"
                cell.fill  = NA_BG
            else:
                cell.fill = row_bg
                if fmt:
                    cell.number_format = fmt

    # ── highlight best value per metric column ────────────────────────────────
    metric_col_keys = [
        (5, "best_val_acc"),
        (6, "test_accuracy"),
        (7, "test_precision"),
        (8, "test_recall"),
        (9, "test_f1"),
    ]
    for ci, key in metric_col_keys:
        candidates = [(ri, r.get(key)) for ri, r in enumerate(comparison, 2) if r.get(key) is not None]
        if candidates:
            best_ri, _ = max(candidates, key=lambda x: x[1])
            best_cell  = ws.cell(row=best_ri, column=ci)
            best_cell.fill = BEST_BG
            best_cell.font = best_font

    # ── footer note ───────────────────────────────────────────────────────────
    note_row = len(comparison) + 3
    note = (
        "Notes: CNN/ResNet/MobileNet — test accuracy evaluated on held-out test set. "
        "YOLO models — metrics from final training epoch (validation split). "
        "Green highlight = best value in column."
    )
    nc = ws.cell(row=note_row, column=1, value=note)
    nc.font      = note_font
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=len(COLS),
    )
    ws.row_dimensions[note_row].height = 32

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def load_model_training_details() -> dict:
    """Load per-model training details from history JSON and YOLO results CSV."""
    details = {}

    for model_name in [m for m in config.CLASSIFY_MODELS if not m.startswith("yolo")]:
        history_path = config.OUTPUT_DIR / f"history_{model_name}.json"
        if not history_path.exists():
            continue
        with open(history_path) as f:
            h = json.load(f)
        train_acc = h.get("train_acc", [])
        val_acc   = h.get("val_acc", [])
        details[model_name] = {
            "epochs":           len(train_acc),
            "final_train_acc":  round(train_acc[-1], 2) if train_acc else None,
            "final_val_acc":    round(val_acc[-1], 2)   if val_acc   else None,
            "best_val_acc":     round(max(val_acc), 2)  if val_acc   else None,
            "final_train_loss": round(h["train_loss"][-1], 4) if h.get("train_loss") else None,
            "final_val_loss":   round(h["val_loss"][-1], 4)   if h.get("val_loss")   else None,
            "total_time_s":     round(sum(h.get("epoch_times", [])), 1),
        }

    yolo_classify_csv = config.YOLO26_CLASSIFY_RUN_DIR / "results.csv"
    if yolo_classify_csv.exists():
        with open(yolo_classify_csv) as f:
            rows = list(csv.DictReader(f))
        if rows:
            last = {k.strip(): v.strip() for k, v in rows[-1].items()}
            details["yolo26_classify"] = {
                "epochs":           int(float(last.get("epoch", len(rows)))),
                "final_val_acc":    round(float(last["metrics/accuracy_top1"]) * 100, 2) if last.get("metrics/accuracy_top1") else None,
                "final_train_loss": round(float(last["train/loss"]), 4)  if last.get("train/loss") else None,
                "final_val_loss":   round(float(last["val/loss"]), 4)    if last.get("val/loss")   else None,
                "total_time_s":     round(float(last["time"]), 1)        if last.get("time")       else None,
            }

    yolo_detect_csv = config.YOLO26_DETECT_RUN_DIR / "results.csv"
    if yolo_detect_csv.exists():
        with open(yolo_detect_csv) as f:
            rows = list(csv.DictReader(f))
        if rows:
            last = {k.strip(): v.strip() for k, v in rows[-1].items()}
            details["yolo26"] = {
                "epochs":      int(float(last.get("epoch", len(rows)))),
                "map50":       round(float(last["metrics/mAP50(B)"]) * 100, 2)    if last.get("metrics/mAP50(B)")    else None,
                "precision":   round(float(last["metrics/precision(B)"]) * 100, 2) if last.get("metrics/precision(B)") else None,
                "recall":      round(float(last["metrics/recall(B)"]) * 100, 2)   if last.get("metrics/recall(B)")   else None,
                "total_time_s": round(float(last["time"]), 1)                      if last.get("time")                else None,
            }

    # Fallback: the live training CSV is regenerated per run and can be absent
    # (e.g. outputs were cleaned). The evaluate-all comparison still holds the
    # YOLO detect mAP@50 (stored as test_accuracy), so surface it from there so
    # the model card keeps showing mAP instead of going blank.
    if "yolo26" not in details:
        comp_path = config.OUTPUT_DIR / "model_comparison.json"
        if comp_path.exists():
            try:
                with open(comp_path) as f:
                    comp = json.load(f)
                entry = next((m for m in comp if m.get("model") == "yolo26"), None)
                if entry and entry.get("test_accuracy") is not None:
                    details["yolo26"] = {
                        "map50":     entry.get("test_accuracy"),
                        "precision": entry.get("test_precision"),
                        "recall":    entry.get("test_recall"),
                    }
            except (json.JSONDecodeError, ValueError, OSError):
                pass

    return details
