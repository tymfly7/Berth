"""
Berth — FastAPI Backend
====================================
REST + WebSocket endpoints for real-time parking detection.

Features:
  - /predict endpoint: upload image, get slot-wise availability
  - WebSocket video streaming at ~20 FPS
  - API key auth (optional via BERTH_API_KEY)
  - Rate limiting on uploads
  - Training management endpoints
  - Model switching (cnn_scratch / resnet50 / mobilenetv4 / yolo26_classify / yolo26)
  - Multi-camera registry with persistent activation
"""

import base64
import hmac
import os
import re
import sys
import json
import uuid
import time
import asyncio
import logging
import threading
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional
from urllib.parse import urlparse
import cv2
import numpy as np
import uvicorn
from fastapi import (
    FastAPI, WebSocket, WebSocketDisconnect,
    UploadFile, File, Form, HTTPException, Request, Depends,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from PIL import Image
import io
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from src.inference.video_processor import VideoProcessor
from src.roi.roi_store import RoiStore
from src.cameras.camera_registry import camera_registry
from src.db import database as db
import config


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
)
logger = logging.getLogger("berth")
sys.path.insert(0, str(Path(__file__).parent))

limiter = Limiter(key_func=get_remote_address)

@asynccontextmanager
async def lifespan(app: FastAPI):
    db.init_db()
    if not config.API_KEY:
        logger.warning(
            "BERTH_API_KEY is not set — all protected endpoints are publicly accessible. "
            "Set this env var before any network-facing deployment."
        )
    from src.sync.sync_worker import SyncWorker
    SyncWorker().start()
    # Restore active cameras and pre-warm the default processor in a background
    # thread so the server starts accepting connections immediately — model loading
    # is slow (5–15 s) and must not block the asyncio event loop.
    def _startup_warmup():
        camera_registry._restore_active()
        try:
            _get_processor()
            logger.info("VideoProcessor pre-warmed")
        except Exception as e:
            logger.warning(f"Processor pre-warm skipped: {e}")
        # Pre-load every classifier so the first analyze of each model isn't
        # delayed by a cold ~2 s load (users switch models to compare them).
        for _name in ("cnn_scratch", "resnet50", "mobilenetv4s", "yolo26_classify", "yolo26"):
            try:
                _get_classifier(_name)
            except Exception as e:
                logger.warning(f"Classifier pre-warm skipped for {_name}: {e}")
        logger.info("Classifiers pre-warmed")
    threading.Thread(target=_startup_warmup, daemon=True, name="startup-warmup").start()
    yield
    camera_registry.shutdown()

app = FastAPI(
    title="Berth",
    description="Real-time parking detection powered by deep learning",
    version="1.0.0",
    lifespan=lifespan,
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    return JSONResponse(
        status_code=404,
        content={"detail": "We looked everywhere and we couldn't find that!"},
    )
_allowed_origins = [o for o in [
    "http://localhost:5173",
    "http://localhost:3000",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:3000",
    os.getenv("BERTH_ALLOWED_ORIGIN", ""),
] if o]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

API_KEY = config.API_KEY

async def verify_api_key(request: Request):
    if not API_KEY:
        return
    key = request.headers.get("X-API-Key", "")
    if not hmac.compare_digest(key.encode(), API_KEY.encode()):
        raise HTTPException(401, "Invalid or missing API key")

# ── Camera source validation (SSRF guard) ────────────────
_YOUTUBE_HOSTS = {"www.youtube.com", "youtube.com", "youtu.be", "m.youtube.com"}

def _validate_camera_source(source: str, type_: str) -> None:
    if type_ == "usb":
        try:
            idx = int(source)
            if idx < 0:
                raise ValueError
        except (ValueError, TypeError):
            raise HTTPException(400, "USB source must be a non-negative integer device index")
    elif type_ == "rtsp":
        p = urlparse(source)
        if p.scheme not in ("rtsp", "rtsps"):
            raise HTTPException(400, "RTSP source must use rtsp:// or rtsps:// scheme")
        if not p.hostname:
            raise HTTPException(400, "RTSP source must include a hostname")
    elif type_ == "youtube":
        p = urlparse(source)
        if p.hostname not in _YOUTUBE_HOSTS:
            raise HTTPException(400, "YouTube source must be a youtube.com or youtu.be URL")

# ── Active-operation tracking ─────────────────────────────
_active_operations: dict = {}
_ops_lock = threading.Lock()

def _register_op(op_type: str, label: str) -> str:
    op_id = uuid.uuid4().hex[:8]
    now = datetime.now(timezone.utc)
    with _ops_lock:
        # Evict stale entries (crashed ops that never called _finish_op).
        stale = [k for k, v in _active_operations.items()
                 if (now - datetime.fromisoformat(v["started_at"])).total_seconds() > 3600]
        for k in stale:
            del _active_operations[k]
        _active_operations[op_id] = {
            "id": op_id, "type": op_type, "label": label,
            "progress": 0.0,
            "started_at": now.isoformat(),
        }
    return op_id

def _update_op_progress(op_id: str, progress: float) -> None:
    with _ops_lock:
        if op_id in _active_operations:
            _active_operations[op_id]["progress"] = progress

def _finish_op(op_id: str) -> None:
    with _ops_lock:
        _active_operations.pop(op_id, None)

# ── Processor (lazy loaded, thread-safe) ─────────────────
_processor = None
_active_mode = config.ACTIVE_MODEL
_processor_lock = threading.Lock()
_anomaly_enabled = False

# ── model_info cache (invalidated on training start / dataset upload) ─
_model_info_cache: dict = {"data": None, "ts": 0.0}
_MODEL_INFO_TTL = 60.0  # seconds

# ── Classifier cache — one loaded instance per model name ─
_clf_cache: dict = {}
_clf_lock = threading.Lock()

def _get_classifier(model_name: str):
    # 'yolo26' and 'yolo26_classify' load the same weights — share one cached
    # instance so the model isn't held in memory twice.
    cache_key = "yolo26_classify" if model_name in ("yolo26", "yolo26_classify") else model_name
    with _clf_lock:
        if cache_key not in _clf_cache:
            from src.inference.classifier import get_classifier
            clf = get_classifier(model_name=cache_key)
            clf.load()
            if not clf.is_loaded():
                raise Exception(f"Model '{model_name}' failed to load")
            _clf_cache[cache_key] = clf
        return _clf_cache[cache_key]

def _clear_clf_cache():
    with _clf_lock:
        _clf_cache.clear()

def _get_processor():
    global _processor
    with _processor_lock:
        if _processor is None:
            try:
                _processor = VideoProcessor(model_name=_active_mode or None)
                logger.info(f"VideoProcessor initialised (model={_active_mode or 'none'})")
            except Exception as e:
                logger.error(f"VideoProcessor failed to initialise: {e}")
                raise
    return _processor

def _reset_processor():
    global _processor
    with _processor_lock:
        if _processor is not None:
            try:
                _processor.stop_processing()
            except Exception:
                pass
        _processor = None
    _clear_clf_cache()

def _resolve_model_name():
    """Resolve active model name for single-image prediction."""
    supported = ("cnn_scratch", "resnet50", "mobilenetv4s", "yolo26_classify", "yolo26")
    if _active_mode in supported:
        return _active_mode
    for name, path in [
        ("yolo26_classify", config.YOLO26_CLASSIFY_PATH),
        ("cnn_scratch", config.CNN_SCRATCH_PATH),
        ("resnet50", config.RESNET50_PATH),
        ("mobilenetv4s", config.MOBILENETV4_PATH),
    ]:
        if path.exists():
            return name
    return None


def _read_image_from_bytes(filename: str, content: bytes) -> np.ndarray:
    allowed = (".jpg", ".jpeg", ".png", ".bmp")
    if not filename.lower().endswith(allowed):
        raise HTTPException(400, "Unsupported image format. Use JPG or PNG.")
    frame = cv2.imdecode(np.frombuffer(content, np.uint8), cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(400, "Could not decode image")
    return frame


async def _read_image(file: UploadFile) -> np.ndarray:
    content = await file.read()
    return _read_image_from_bytes(file.filename, content)


def _frame_to_b64(frame: np.ndarray) -> str:
    _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 90])
    return base64.b64encode(buf).decode("utf-8")

# ═══════════════════════════════════════════════════════════════
# REST Endpoints
# ═══════════════════════════════════════════════════════════════

@app.get("/")
def root():
    return {
        "service": "Berth",
        "version": "1.0.0",
        "status": "running",
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }

@app.get("/api/health")
def health():
    proc = _get_processor()
    return {
        "status": "ok",
        "processor": type(proc).__name__,
        "model": _active_mode,
        "auth_enabled": bool(API_KEY),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }

@app.get("/api/status")
def get_status():
    with _ops_lock:
        ops = list(_active_operations.values())
    return {"busy": len(ops) > 0, "operations": ops}

# ── Predict endpoint (single spot) ───────────────────────
@app.post("/api/predict", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def predict(request: Request, file: UploadFile = File(...)):
    """
    Upload a parking space image and classify as occupied/vacant.
    Works best with cropped images of individual parking spots.
    """
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(413, "Image exceeds 20 MB limit")

    # Validate extension and decode via shared helper
    frame = await _read_image_from_bytes(file.filename, content)
    pil_image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))

    model_name = _resolve_model_name()
    if model_name is None:
        raise HTTPException(400, "No trained model available. Train a model first.")
    try:
        clf = _get_classifier(model_name)
        result = clf.predict(pil_image)
        result["model"] = model_name
        result["type"] = "single_spot"
    except Exception as e:
        logger.error(f"Prediction error: {e}")
        result = {"status": "error", "message": str(e), "confidence": 0.0}

    return result


# ── Analyze full parking lot image ───────────────────────
@app.post("/api/analyze-lot", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def analyze_lot(request: Request, file: UploadFile = File(...),
                      rows: int = 3, cols: int = 6):
    """
    Upload ANY parking lot image (aerial, Google, etc.) and analyze it.

    Splits the image into a grid of (rows x cols) cells, classifies each
    cell as occupied/vacant, and returns:
    - Slot-wise results with confidence
    - Summary stats (total, available, occupied)
    - Base64-encoded annotated image with green/red overlays

    Query params:
    - rows: number of grid rows (default: 3)
    - cols: number of grid columns (default: 6)
    """
    op_id = _register_op("analysis", "Analyzing parking lot…")
    try:
        if not 1 <= rows <= 50 or not 1 <= cols <= 50:
            raise HTTPException(400, "rows and cols must each be between 1 and 50")

        frame = await _read_image(file)

        model_name = _resolve_model_name()
        if model_name is None:
            raise HTTPException(400, "No trained model available. Train a model first.")
        try:
            clf = _get_classifier(model_name)
        except Exception as e:
            raise HTTPException(400, str(e))

        h, w = frame.shape[:2]
        cell_h = h // rows
        cell_w = w // cols

        cells = []
        slot_id = 1
        for r in range(rows):
            for c in range(cols):
                y1 = r * cell_h
                y2 = (r + 1) * cell_h if r < rows - 1 else h
                x1 = c * cell_w
                x2 = (c + 1) * cell_w if c < cols - 1 else w
                cells.append((slot_id, frame[y1:y2, x1:x2], x1, y1, x2, y2))
                slot_id += 1

        predictions = clf.predict_batch([c[1] for c in cells])

        annotated = frame.copy()
        slots = []
        available = 0
        occupied = 0
        total_conf = 0.0

        for (sid, _, x1, y1, x2, y2), pred in zip(cells, predictions):
            status = pred["status"]
            conf = pred["confidence"]
            total_conf += conf

            if status == "vacant":
                available += 1
                color = (0, 200, 0)
                overlay_color = (0, 255, 0)
            else:
                occupied += 1
                color = (0, 0, 220)
                overlay_color = (0, 0, 255)

            overlay = annotated.copy()
            cv2.rectangle(overlay, (x1, y1), (x2, y2), overlay_color, -1)
            cv2.addWeighted(overlay, 0.25, annotated, 0.75, 0, annotated)
            cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)
            label = f"#{sid} {status[:3].upper()} {conf:.0%}"
            font_scale = max(0.35, min(cell_w, cell_h) / 200)
            cv2.putText(annotated, label, (x1 + 4, y1 + 18),
                        cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), 1)

            slots.append({
                "id": sid,
                "status": status,
                "confidence": round(conf, 4),
                "bbox": f"{x1} {y1} {x2-x1} {y2-y1}",
            })

        total = len(slots)
        avg_conf = total_conf / total if total > 0 else 0
        occ_pct = round(100.0 * occupied / total, 1) if total > 0 else 0

        return {
            "type": "lot_analysis",
            "model": _active_mode,
            "grid": f"{rows}x{cols}",
            "total": total,
            "available": available,
            "occupied": occupied,
            "occupancy_percent": occ_pct,
            "avg_confidence": round(avg_conf, 4),
            "slots": slots,
            "annotated_image": _frame_to_b64(annotated),
        }
    finally:
        _finish_op(op_id)


@app.post("/api/analyze-roi", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def analyze_roi(
    request: Request,
    file: UploadFile = File(...),
    camera_id: str = "default",
    model_name: str = None,
    rois_json: Optional[str] = Form(default=None),
):
    """
    Analyze a parking lot image using saved ROI polygons for the given camera.
    Each ROI polygon is classified as occupied/vacant using the active model.

    Pass rois_json (JSON-encoded list) to bypass the disk read and eliminate
    the sequential save→analyze round-trip from the client.
    """
    op_id = _register_op("roi_analysis", "Analyzing with ROIs…")
    try:
        frame = await _read_image(file)

        # Prefer inline ROIs from the request body (eliminates save→analyze round-trip)
        rois = None
        if rois_json:
            try:
                rois = json.loads(rois_json) or None
            except (json.JSONDecodeError, ValueError):
                pass
        if not rois:
            rois = RoiStore.get_rois(camera_id)
        if not rois:
            raise HTTPException(400, "No ROIs saved for this camera. Draw and save ROIs first.")

        supported = ("cnn_scratch", "resnet50", "mobilenetv4s", "yolo26_classify", "yolo26")
        model_name = model_name if model_name in supported else _resolve_model_name()
        if model_name is None:
            raise HTTPException(400, "No trained model available. Train a model first.")
        try:
            clf = _get_classifier(model_name)
        except Exception as e:
            raise HTTPException(400, str(e))

        h, w = frame.shape[:2]
        annotated = frame.copy()
        available = 0
        occupied = 0
        total_conf = 0.0

        # Pass 1 — filter valid ROIs and collect crops (no inference yet)
        valid = []
        for roi in rois:
            polygon = roi.get("polygon", [])
            if len(polygon) < 3:
                continue
            xs = [max(0, min(w - 1, int(p[0] * w))) for p in polygon]
            ys = [max(0, min(h - 1, int(p[1] * h))) for p in polygon]
            x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
            if x2 <= x1 or y2 <= y1:
                continue
            valid.append({"roi": roi, "crop": frame[y1:y2, x1:x2],
                          "xs": xs, "ys": ys, "polygon": polygon})

        # Single batched forward pass — all N crops classified in one model call
        predictions = clf.predict_batch([v["crop"] for v in valid]) if valid else []

        # Pass 2 — draw all semi-transparent fills in one blend (avoids N full-image copies)
        overlay = annotated.copy()
        for entry, pred in zip(valid, predictions):
            pts = np.array([[int(p[0] * w), int(p[1] * h)]
                            for p in entry["polygon"]], np.int32)
            overlay_color = (0, 255, 0) if pred["status"] == "vacant" else (0, 0, 255)
            cv2.fillPoly(overlay, [pts], overlay_color)
        cv2.addWeighted(overlay, 0.3, annotated, 0.7, 0, annotated)

        # Pass 3 — outlines, labels, slot data
        slots = []
        for entry, pred in zip(valid, predictions):
            roi = entry["roi"]
            xs, ys = entry["xs"], entry["ys"]
            status = pred["status"]
            conf = pred["confidence"]
            total_conf += conf

            if status == "vacant":
                available += 1
                color = (0, 200, 0)
            else:
                occupied += 1
                color = (0, 0, 220)

            pts = np.array([[int(p[0] * w), int(p[1] * h)]
                            for p in entry["polygon"]], np.int32)
            cv2.polylines(annotated, [pts], True, color, 2)
            cx, cy = sum(xs) // len(xs), sum(ys) // len(ys)
            cv2.putText(annotated,
                        f"{roi.get('label', 'Slot')} {status[:3].upper()} {conf:.0%}",
                        (cx - 20, cy), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
            slots.append({
                "id": roi.get("id"),
                "label": roi.get("label", "Slot"),
                "status": status,
                "confidence": round(conf, 4),
            })

        total = len(slots)
        occ_pct = round(100.0 * occupied / total, 1) if total > 0 else 0
        avg_conf = round(total_conf / total, 4) if total > 0 else 0

        # Cap output resolution — the result is shown in a narrow side panel, so a
        # full-res annotated image only bloats the base64 response payload.
        out_max = max(h, w)
        if out_max > 1280:
            s = 1280 / out_max
            annotated = cv2.resize(annotated, (int(w * s), int(h * s)),
                                   interpolation=cv2.INTER_AREA)
        annotated_b64 = _frame_to_b64(annotated)

        return {
            "type": "roi_analysis",
            "model": _active_mode,
            "total": total,
            "available": available,
            "occupied": occupied,
            "occupancy_percent": occ_pct,
            "avg_confidence": avg_conf,
            "slots": slots,
            "annotated_image": annotated_b64,
        }
    finally:
        _finish_op(op_id)


# ── Analyze misparked vehicles (YOLO + ROI geometry) ────
@app.post("/api/analyze-misparked", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def analyze_misparked(
    request: Request,
    file: UploadFile = File(...),
    camera_id: str = "default",
):
    """
    Detect misparked vehicles in an uploaded parking lot image.

    Runs YOLO26 to locate all vehicles, loads the camera's ROI polygons,
    then classifies each vehicle as ok / straddling / outside_markings.

    Returns JSON with per-slot occupancy, a misparked-vehicle list, a
    misparked count (int) for the metrics shape, and a base64-encoded
    annotated image with misparked cars highlighted in orange.
    """
    op_id = _register_op("misparked_analysis", "Detecting misparked vehicles…")
    try:
        frame = await _read_image(file)

        rois = RoiStore.get_rois(camera_id)
        if not rois:
            raise HTTPException(
                400,
                f"No ROIs saved for camera '{camera_id}'. Draw and save ROIs first.",
            )

        # Load YOLO26 detector — surface a clear error if weights are missing
        try:
            from src.models.yolo_detector import ParkingYOLO26
            detector = ParkingYOLO26(str(config.YOLO26_DETECT_PATH))
        except FileNotFoundError:
            raise HTTPException(
                400,
                "YOLO26 detector weights not found. "
                "Train a YOLO26 detector first via the Training panel.",
            )
        except RuntimeError as exc:
            logger.error(f"YOLO26 detector failed to load: {exc}")
            raise HTTPException(400, f"YOLO26 detector failed to load: {exc}")

        h, w = frame.shape[:2]
        detections = detector.predict_frame(frame)
        cars = [{"bbox": d["bbox"], "confidence": d["confidence"]} for d in detections]

        from src.inference.parking_geometry import aggregate_lot, classify_vehicle_parking
        lot = aggregate_lot(cars, rois, w, h)

        # ── Annotate image ────────────────────────────────────────────────────
        annotated = frame.copy()

        # Draw slot outlines colour-coded by occupancy
        for slot in lot["slots"]:
            roi = next((r for r in rois if r.get("id") == slot["id"]), None)
            if roi is None:
                continue
            polygon = roi.get("polygon", [])
            if len(polygon) < 3:
                continue
            pts = np.array(
                [[int(p[0] * w), int(p[1] * h)] for p in polygon], np.int32
            )
            slot_color = (0, 200, 0) if slot["status"] == "vacant" else (0, 0, 200)
            overlay = annotated.copy()
            cv2.fillPoly(overlay, [pts], slot_color)
            cv2.addWeighted(overlay, 0.20, annotated, 0.80, 0, annotated)
            cv2.polylines(annotated, [pts], True, slot_color, 2)
            xs = [int(p[0] * w) for p in polygon]
            ys = [int(p[1] * h) for p in polygon]
            cx, cy = sum(xs) // len(xs), sum(ys) // len(ys)
            cv2.putText(
                annotated,
                slot.get("label", ""),
                (cx - 10, cy),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.4,
                (255, 255, 255),
                1,
            )

        # Draw each detected car — orange for misparked, light green for ok
        for car in cars:
            clf = classify_vehicle_parking(car["bbox"], rois, w, h)
            x1, y1, x2, y2 = (int(v) for v in car["bbox"])
            if clf["status"] == "misparked":
                color = (0, 165, 255)   # BGR orange
                reason_short = "STRADDLE" if clf["reason"] == "straddling" else "OUTSIDE"
                label = f"MISPK {reason_short}"
            else:
                color = (144, 238, 144)  # light green
                label = f"OK {car['confidence']:.0%}"
            cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)
            cv2.putText(
                annotated,
                label,
                (x1 + 2, max(y1 - 4, 14)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.45,
                color,
                1,
            )

        return {
            "type": "misparked_analysis",
            "camera_id": camera_id,
            "total": lot["total"],
            "available": lot["available"],
            "occupied": lot["occupied"],
            "misparked": lot["misparked_count"],        # int for metrics shape
            "misparked_vehicles": lot["misparked"],     # detailed list
            "slots": lot["slots"],
            "annotated_image": _frame_to_b64(annotated),
        }
    finally:
        _finish_op(op_id)


# ── Public metrics (no auth) ─────────────────────────────
@app.get("/api/public/metrics")
def get_public_metrics():
    # Aggregate across active cameras (mirrors /api/history); fall back to the
    # default processor when no cameras are active.
    active_procs = [
        camera_registry.get_processor(c["id"])
        for c in camera_registry.get_all()
        if c.get("active")
    ]
    active_procs = [p for p in active_procs if p is not None]
    if not active_procs:
        return _get_processor().get_metrics()

    metrics = [p.get_metrics() for p in active_procs]
    total     = sum(m.get("total", 0)     for m in metrics)
    available = sum(m.get("available", 0) for m in metrics)
    occupied  = sum(m.get("occupied", 0)  for m in metrics)
    return {
        **metrics[0],
        "total": total,
        "available": available,
        "occupied": occupied,
        "occupancy_percent": round(100.0 * occupied / total, 1) if total else 0.0,
        "avg_confidence": round(sum(m.get("avg_confidence", 0.0) for m in metrics) / len(metrics), 4),
        "fps": round(sum(m.get("fps", 0.0) for m in metrics) / len(metrics), 1),
        "misparked_count": sum(m.get("misparked_count", 0) for m in metrics),
        "anomaly_enabled": any(m.get("anomaly_enabled") for m in metrics),
        "slots": [s for m in metrics for s in m.get("slots", [])],
    }

# ── Metrics / Heatmap / History ──────────────────────────
@app.get("/api/metrics", dependencies=[Depends(verify_api_key)])
def get_metrics():
    return _get_processor().get_metrics()

@app.get("/api/heatmap", dependencies=[Depends(verify_api_key)])
def get_heatmap():
    active = next((c for c in camera_registry.get_all() if c.get("active")), None)
    if active:
        proc = camera_registry.get_processor(active["id"])
        if proc and hasattr(proc, "get_heatmap"):
            return proc.get_heatmap()
    proc = _get_processor()
    return proc.get_heatmap() if hasattr(proc, "get_heatmap") else []

@app.get("/api/heatmap/{camera_id}", dependencies=[Depends(verify_api_key)])
def get_heatmap_camera(camera_id: str):
    proc = camera_registry.get_processor(camera_id)
    if proc and hasattr(proc, "get_heatmap"):
        return proc.get_heatmap()
    return []

@app.get("/api/history", dependencies=[Depends(verify_api_key)])
def get_history():
    # Prefer active camera processors; fall back to the default processor
    active_procs = [
        camera_registry.get_processor(c["id"])
        for c in camera_registry.get_all()
        if c.get("active")
    ]
    active_procs = [p for p in active_procs if p and hasattr(p, "get_history")]
    if active_procs:
        # Merge and sort all camera histories by timestamp
        merged = sorted(
            (entry for p in active_procs for entry in p.get_history()),
            key=lambda e: e.get("timestamp", "")
        )
        return merged[-100:]
    proc = _get_processor()
    return proc.get_history() if hasattr(proc, "get_history") else []

@app.get("/api/trends", dependencies=[Depends(verify_api_key)])
def get_trends(range: str = "day", camera_id: str = None):
    if range not in ("today", "day", "week", "month"):
        raise HTTPException(400, "range must be today, day, week, or month")
    return db.query_trends(range, camera_id)

@app.get("/api/alerts", dependencies=[Depends(verify_api_key)])
def get_alerts(limit: int = 50):
    return db.get_alerts(limit)

@app.get("/api/training-runs", dependencies=[Depends(verify_api_key)])
def get_training_runs(limit: int = 20):
    return db.get_training_runs(limit)

# ── Video / Camera ───────────────────────────────────────
@app.post("/api/upload-video", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def upload_video(request: Request, file: UploadFile = File(...)):
    op_id = _register_op("video_upload", "Uploading video…")
    try:
        allowed = (".mp4", ".avi", ".mov", ".mkv", ".webm")
        if not file.filename.lower().endswith(allowed):
            raise HTTPException(400, "Unsupported video format")
        safe_name = Path(file.filename).name  # strip any directory components
        dest = config.UPLOAD_DIR / safe_name
        max_bytes = 500 * 1024 * 1024  # 500 MB
        written = 0
        try:
            with open(dest, "wb") as f_out:
                while chunk := await file.read(1 << 20):
                    written += len(chunk)
                    if written > max_bytes:
                        raise HTTPException(413, "Video exceeds 500 MB limit")
                    f_out.write(chunk)
        except HTTPException:
            dest.unlink(missing_ok=True)
            raise
        proc = _get_processor()
        proc.set_video_source(str(dest))
        return {"message": "Video uploaded", "path": safe_name}
    finally:
        _finish_op(op_id)

@app.post("/api/use-camera", dependencies=[Depends(verify_api_key)])
def use_camera():
    proc = _get_processor()
    proc.set_video_source(0)
    return {"message": "Switched to live camera"}

# ── Anomaly detection settings ────────────────────────────
@app.get("/api/settings/anomaly", dependencies=[Depends(verify_api_key)])
def get_anomaly():
    return {"enabled": _anomaly_enabled}

@app.post("/api/settings/anomaly", dependencies=[Depends(verify_api_key)])
async def set_anomaly(request: Request):
    global _anomaly_enabled
    body = await request.json()
    enabled = bool(body.get("enabled", False))
    try:
        _get_processor().set_anomaly_detection(enabled)
    except FileNotFoundError:
        raise HTTPException(400, "YOLO26 detect model not found. Train it first via the Training panel.")
    except RuntimeError as e:
        raise HTTPException(400, str(e))
    _anomaly_enabled = enabled
    for cam in camera_registry.get_all():
        if cam.get("active"):
            p = camera_registry.get_processor(cam["id"])
            if p:
                try:
                    p.set_anomaly_detection(enabled)
                except Exception as e:
                    logger.debug(f"Anomaly toggle skipped for camera {cam['id']}: {e}")
    return {"enabled": _anomaly_enabled}

# ── Model switching ──────────────────────────────────────
@app.post("/api/use-model/{model_name}", dependencies=[Depends(verify_api_key)])
def use_model(model_name: str):
    global _active_mode
    valid = ["cnn_scratch", "resnet50", "mobilenetv4s", "yolo26_classify", "yolo26"]
    if model_name not in valid:
        raise HTTPException(400, f"Invalid model. Choose from: {valid}")
    _reset_processor()
    _active_mode = model_name
    proc = _get_processor()
    proc.start_processing()
    # Restart all active live cameras with the new model so they pick it up immediately.
    restarted = 0
    for cam in camera_registry.get_all():
        if cam.get("active"):
            camera_registry.activate(cam["id"], model_name=model_name)
            restarted += 1
    return {"message": f"Switched to {model_name}", "cameras_restarted": restarted}

@app.post("/api/test-model/{model_name}", dependencies=[Depends(verify_api_key)])
def test_model(model_name: str):
    if model_name in ("yolo26", "yolo26_detect"):
        raise HTTPException(400, "YOLO26 detect uses a detection interface — per-patch accuracy testing is not supported.")
    testable = ["cnn_scratch", "resnet50", "mobilenetv4s", "yolo26_classify"]
    if model_name not in testable:
        raise HTTPException(400, f"Unknown model '{model_name}'. Testable: {testable}")
    try:
        import torch
        from src.models.model_factory import load_model
        from src.data_prep.preprocessor import prepare_dataset
        from src.eval.evaluator import evaluate_model
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        model = load_model(model_name, device=device)
        data = prepare_dataset()
        test_loader = data["test_loader"]
        metrics = evaluate_model(model, test_loader, device=device)
        return {"model": model_name, **metrics}
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        raise HTTPException(500, f"Test failed: {e}")

@app.post("/api/evaluate/all", dependencies=[Depends(verify_api_key)])
def evaluate_all():
    if config.DEPLOYMENT_PROFILE == "edge":
        raise HTTPException(403, "Evaluation is not available on edge nodes. Use the hub server.")
    from src.train.train_manager import TrainManager
    result = TrainManager().start_evaluation()
    if result.get("status") == "error":
        raise HTTPException(400, result["message"])
    return result


def _build_comparison_excel(comparison: list) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model Comparison"

    # ── palette ───────────────────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill(fgColor=hex_color, fill_type="solid")

    HDR_BG   = fill("3730a3")   # indigo-700
    CNN_BG   = fill("dbeafe")   # blue-100
    YCLS_BG  = fill("fef3c7")   # amber-100
    YDET_BG  = fill("ede9fe")   # violet-100
    BEST_BG  = fill("bbf7d0")   # green-200
    NA_BG    = fill("f1f5f9")   # slate-100

    thin   = Side(style="thin",   color="94a3b8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    hdr_font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    best_font = Font(bold=True, size=10, name="Calibri")
    note_font = Font(italic=True, size=9, color="64748b", name="Calibri")

    # ── column definitions ────────────────────────────────────────────────────
    # (header text, width, data key, format)
    COLS = [
        ("Model",            20, "model",            None),
        ("Type",             14, "type",              None),
        ("Epochs",            8, "epochs",            "0"),
        ("Train Time (s)",   12, "train_time",        "0.0"),
        ("Best Val Acc (%)", 13, "best_val_acc",      "0.00"),
        ("Test Acc (%)",     12, "test_accuracy",     "0.00"),
        ("Precision (%)",    12, "test_precision",    "0.00"),
        ("Recall (%)",       12, "test_recall",       "0.00"),
        ("F1 Score (%)",     12, "test_f1",           "0.00"),
        ("Total Params",     13, "total_params",      "#,##0"),
        ("Trainable Params", 14, "trainable_params",  "#,##0"),
    ]

    MODEL_LABELS = {
        "cnn_scratch":     "CNN Scratch",
        "resnet50":        "ResNet-50",
        "mobilenetv4s":     "MobileNetV4",
        "yolo26_classify": "YOLO26 Classify",
        "yolo26":          "YOLO26 Detect",
    }
    ROW_BG = {
        "cnn_scratch":     CNN_BG,
        "resnet50":        CNN_BG,
        "mobilenetv4s":     CNN_BG,
        "yolo26_classify": YCLS_BG,
        "yolo26":          YDET_BG,
    }

    from openpyxl.utils import get_column_letter

    # ── header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    for ci, (hdr, width, _, _) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill      = HDR_BG
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── data rows ─────────────────────────────────────────────────────────────
    for ri, r in enumerate(comparison, 2):
        name   = r.get("model", "")
        row_bg = ROW_BG.get(name, fill("FFFFFF"))
        ws.row_dimensions[ri].height = 20

        for ci, (_, _, key, fmt) in enumerate(COLS, 1):
            if key == "model":
                val = MODEL_LABELS.get(name, name)
                align = left
            elif key == "type":
                val   = str(r.get("type", "classification")).title()
                align = center
            else:
                raw = r.get(key)
                val = round(raw, 2) if isinstance(raw, float) else raw
                align = center

            cell = ws.cell(row=ri, column=ci, value=val if val is not None else None)
            cell.font      = body_font
            cell.alignment = align
            cell.border    = border

            if val is None:
                cell.value = "—"
                cell.fill  = NA_BG
            else:
                cell.fill = row_bg
                if fmt:
                    cell.number_format = fmt

    # ── highlight best value per metric column ────────────────────────────────
    metric_col_keys = [
        (5, "best_val_acc"),
        (6, "test_accuracy"),
        (7, "test_precision"),
        (8, "test_recall"),
        (9, "test_f1"),
    ]
    for ci, key in metric_col_keys:
        candidates = [(ri, r.get(key)) for ri, r in enumerate(comparison, 2) if r.get(key) is not None]
        if candidates:
            best_ri, _ = max(candidates, key=lambda x: x[1])
            best_cell  = ws.cell(row=best_ri, column=ci)
            best_cell.fill = BEST_BG
            best_cell.font = best_font

    # ── footer note ───────────────────────────────────────────────────────────
    note_row = len(comparison) + 3
    note = (
        "Notes: CNN/ResNet/MobileNet — test accuracy evaluated on held-out test set. "
        "YOLO models — metrics from final training epoch (validation split). "
        "Green highlight = best value in column."
    )
    nc = ws.cell(row=note_row, column=1, value=note)
    nc.font      = note_font
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row,   end_column=len(COLS),
    )
    ws.row_dimensions[note_row].height = 32

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.get("/api/evaluate/excel", dependencies=[Depends(verify_api_key)])
def download_evaluation_excel():
    comparison_path = config.OUTPUT_DIR / "model_comparison.json"
    if not comparison_path.exists():
        raise HTTPException(404, "No evaluation results found. Run 'Evaluate All' first.")
    with open(comparison_path) as f:
        comparison = json.load(f)
    xlsx_bytes = _build_comparison_excel(comparison)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=model_comparison.xlsx"},
    )


def _load_model_training_details() -> dict:
    """Load per-model training details from history JSON and YOLO results CSV."""
    import csv
    details = {}

    for model_name in ("cnn_scratch", "resnet50", "mobilenetv4s"):
        history_path = config.OUTPUT_DIR / f"history_{model_name}.json"
        if not history_path.exists():
            continue
        with open(history_path) as f:
            h = json.load(f)
        train_acc = h.get("train_acc", [])
        val_acc   = h.get("val_acc", [])
        details[model_name] = {
            "epochs":           len(train_acc),
            "final_train_acc":  round(train_acc[-1], 2) if train_acc else None,
            "final_val_acc":    round(val_acc[-1], 2)   if val_acc   else None,
            "best_val_acc":     round(max(val_acc), 2)  if val_acc   else None,
            "final_train_loss": round(h["train_loss"][-1], 4) if h.get("train_loss") else None,
            "final_val_loss":   round(h["val_loss"][-1], 4)   if h.get("val_loss")   else None,
            "total_time_s":     round(sum(h.get("epoch_times", [])), 1),
        }

    yolo_classify_csv = config.OUTPUT_DIR / "yolo26_classify" / "run" / "results.csv"
    if yolo_classify_csv.exists():
        with open(yolo_classify_csv) as f:
            rows = list(csv.DictReader(f))
        if rows:
            last = {k.strip(): v.strip() for k, v in rows[-1].items()}
            details["yolo26_classify"] = {
                "epochs":           int(float(last.get("epoch", len(rows)))),
                "final_val_acc":    round(float(last["metrics/accuracy_top1"]) * 100, 2) if last.get("metrics/accuracy_top1") else None,
                "final_train_loss": round(float(last["train/loss"]), 4)  if last.get("train/loss") else None,
                "final_val_loss":   round(float(last["val/loss"]), 4)    if last.get("val/loss")   else None,
                "total_time_s":     round(float(last["time"]), 1)        if last.get("time")       else None,
            }

    yolo_detect_csv = config.OUTPUT_DIR / "yolo26_detect" / "run" / "results.csv"
    if yolo_detect_csv.exists():
        with open(yolo_detect_csv) as f:
            rows = list(csv.DictReader(f))
        if rows:
            last = {k.strip(): v.strip() for k, v in rows[-1].items()}
            details["yolo26"] = {
                "epochs":      int(float(last.get("epoch", len(rows)))),
                "map50":       round(float(last["metrics/mAP50(B)"]) * 100, 2)    if last.get("metrics/mAP50(B)")    else None,
                "precision":   round(float(last["metrics/precision(B)"]) * 100, 2) if last.get("metrics/precision(B)") else None,
                "recall":      round(float(last["metrics/recall(B)"]) * 100, 2)   if last.get("metrics/recall(B)")   else None,
                "total_time_s": round(float(last["time"]), 1)                      if last.get("time")                else None,
            }

    return details


@app.get("/api/model/info", dependencies=[Depends(verify_api_key)])
def model_info():
    now = time.monotonic()
    # Return cached result if still fresh; active_model can change so check it too.
    cached = _model_info_cache["data"]
    if cached and (now - _model_info_cache["ts"]) < _MODEL_INFO_TTL and cached.get("active_model") == _active_mode:
        return cached

    data_dir = config.DATA_DIR
    occ_dir = data_dir / "occupied"
    vac_dir = data_dir / "vacant"
    dataset_ready = occ_dir.exists() and vac_dir.exists()
    occupied_count = len(list(occ_dir.glob("*.*"))) if occ_dir.exists() else 0
    vacant_count = len(list(vac_dir.glob("*.*"))) if vac_dir.exists() else 0
    dataset_count = occupied_count + vacant_count

    comparison_path = config.OUTPUT_DIR / "model_comparison.json"
    comparison = None
    if comparison_path.exists():
        with open(comparison_path) as f:
            comparison = json.load(f)

    result = {
        "active_model": _active_mode,
        "available_models": {
            "cnn_scratch":     config.CNN_SCRATCH_PATH.exists(),
            "resnet50":        config.RESNET50_PATH.exists(),
            "mobilenetv4s":     config.MOBILENETV4_PATH.exists(),
            "yolo26_classify": config.YOLO26_CLASSIFY_PATH.exists(),
            "yolo26":          config.YOLO26_DETECT_PATH.exists(),
        },
        "dataset_ready": dataset_ready,
        "dataset_count": dataset_count,
        "occupied_count": occupied_count,
        "vacant_count": vacant_count,
        "comparison": comparison,
        "model_details": _load_model_training_details(),
    }
    _model_info_cache["data"] = result
    _model_info_cache["ts"] = now
    return result

# ── Training ─────────────────────────────────────────────
@app.post("/api/train/start", dependencies=[Depends(verify_api_key)])
@limiter.limit("20/hour")
def start_training(request: Request, model_name: str = "cnn_scratch",
                   compare_all: bool = False):
    if config.DEPLOYMENT_PROFILE == "edge":
        raise HTTPException(403, "Training is not available on edge nodes. Use the hub server.")
    valid_models = ["cnn_scratch", "resnet50", "mobilenetv4s", "yolo26_classify", "yolo26_detect"]
    if model_name not in valid_models:
        raise HTTPException(400, f"Unknown model '{model_name}'. Choose from: {valid_models}")
    from src.train.train_manager import TrainManager
    mgr = TrainManager()
    if mgr.is_training():
        raise HTTPException(409, "Training already in progress")
    # YOLO detect uses the gopro annotated dataset, not the occupied/vacant folders
    if model_name not in ("yolo26_classify", "yolo26_detect"):
        occ = config.DATA_DIR / "occupied"
        vac = config.DATA_DIR / "vacant"
        if not occ.exists() or not vac.exists():
            raise HTTPException(400, "Dataset not found. Prepare it first.")
    if model_name == "yolo26_detect" and not config.YOLO_GOPRO_DIR.exists():
        raise HTTPException(400, "Gopro annotated dataset not found. Expected: backend/data/yolo_data/parking_rois_gopro/")
    _model_info_cache["data"] = None  # invalidate so next poll reflects new state
    result = mgr.start_training(model_name, compare_all=compare_all)
    op_id = _register_op("training", f"Training {model_name}…")

    def _monitor():
        from src.train.train_manager import TrainManager as TM
        while True:
            time.sleep(2)
            try:
                s = TM().get_status()
                epoch = s.get("epoch") or 0
                total = s.get("total_epochs") or 0
                _update_op_progress(op_id, epoch / total if total > 0 else 0)
                if s.get("status") in ("done", "error", "idle"):
                    break
            except Exception:
                break
        _finish_op(op_id)

    threading.Thread(target=_monitor, daemon=True).start()
    return result

@app.get("/api/train/status", dependencies=[Depends(verify_api_key)])
def train_status():
    from src.train.train_manager import TrainManager
    return TrainManager().get_status()

@app.post("/api/dataset/upload", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def upload_dataset_images(
    request: Request,
    files: List[UploadFile] = File(...),
    label: str = Form(...),
):
    if config.DEPLOYMENT_PROFILE == "edge":
        raise HTTPException(403, "Dataset upload is not available on edge nodes. Use the hub server.")
    op_id = _register_op("dataset_upload", "Saving training images…")
    try:
        if label not in ("occupied", "vacant"):
            raise HTTPException(400, "label must be 'occupied' or 'vacant'")
        if len(files) > 50:
            raise HTTPException(400, "Maximum 50 files per request")

        allowed = {".jpg", ".jpeg", ".png", ".bmp"}
        dest_dir = config.DATA_DIR / label
        dest_dir.mkdir(parents=True, exist_ok=True)

        max_image_bytes = 20 * 1024 * 1024  # 20 MB per image
        saved = 0
        skipped = 0
        for file in files:
            safe_name = Path(file.filename).name  # strip any directory components
            ext = Path(safe_name).suffix.lower()
            if ext not in allowed:
                skipped += 1
                continue
            content = await file.read()
            if len(content) > max_image_bytes:
                skipped += 1
                continue
            dest = dest_dir / safe_name
            if dest.exists():
                stem = Path(safe_name).stem
                suffix_str = uuid.uuid4().hex[:6]
                dest = dest_dir / f"{stem}_{suffix_str}{ext}"
            with open(dest, "wb") as f:
                f.write(content)
            saved += 1

        _model_info_cache["data"] = None  # dataset count changed
        return {"saved": saved, "skipped": skipped, "label": label}
    finally:
        _finish_op(op_id)


@app.post("/api/dataset/prepare", dependencies=[Depends(verify_api_key)])
def prepare_dataset(source: str = None, max_per_class: int = 0,
                    generate_sample: bool = False, sample_count: int = 200):
    from src.data_prep.downloader import organize_pklot, generate_sample_dataset
    if generate_sample:
        generate_sample_dataset(num_per_class=sample_count)
        return {"message": f"Generated {sample_count} synthetic images per class"}
    result = organize_pklot(source_root=source, max_per_class=max_per_class)
    return {"message": "Dataset prepared", **result}

# ── ROI endpoints ────────────────────────────────────────
@app.get("/api/roi/{camera_id}", dependencies=[Depends(verify_api_key)])
def get_rois(camera_id: str):
    return RoiStore.get_rois(camera_id)

@app.get("/api/roi/{camera_id}/snapshot", dependencies=[Depends(verify_api_key)])
def get_snapshot(camera_id: str):
    snap_path = RoiStore.get_snapshot_path(camera_id)
    if snap_path is None:
        raise HTTPException(404, "No snapshot found for this camera")
    return Response(content=snap_path.read_bytes(), media_type="image/jpeg")

@app.post("/api/roi/{camera_id}/snapshot", dependencies=[Depends(verify_api_key)])
async def save_snapshot(camera_id: str, file: UploadFile = File(...)):
    allowed = (".jpg", ".jpeg", ".png")
    if not file.filename.lower().endswith(allowed):
        raise HTTPException(400, "Only JPG and PNG images are supported")
    content = await file.read()
    try:
        return RoiStore.save_snapshot(camera_id, content)
    except ValueError as e:
        raise HTTPException(400, str(e))

@app.post("/api/roi/{camera_id}", dependencies=[Depends(verify_api_key)])
async def save_rois(camera_id: str, request: Request):
    body = await request.json()
    rois = body.get("rois", [])
    try:
        RoiStore.save_rois(camera_id, rois)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"saved": len(rois)}

@app.delete("/api/roi/{camera_id}/{roi_id}", dependencies=[Depends(verify_api_key)])
def delete_roi(camera_id: str, roi_id: str):
    if not RoiStore.delete_roi(camera_id, roi_id):
        raise HTTPException(404, f"ROI '{roi_id}' not found")
    return {"deleted": roi_id}


@app.delete("/api/roi/{camera_id}", dependencies=[Depends(verify_api_key)])
def delete_roi_config(camera_id: str):
    """Delete all ROIs and snapshot for a camera/lot config."""
    roi_path = RoiStore._roi_path(camera_id)
    snap_path = RoiStore._snapshot_path(camera_id)
    if not roi_path.exists():
        raise HTTPException(404, f"No ROI config found for '{camera_id}'")
    roi_path.unlink()
    if snap_path.exists():
        snap_path.unlink()
    return {"deleted": camera_id}


@app.post("/api/roi/{camera_id}/propose", dependencies=[Depends(verify_api_key)])
@limiter.limit(config.UPLOAD_RATE_LIMIT)
async def propose_rois(
    request: Request,
    camera_id: str,
    use_line_detection: bool = False,
    file: Optional[UploadFile] = File(default=None),
):
    """
    Auto-detect candidate parking-spot ROIs from an image.

    Accepts an uploaded image (multipart form field 'file'), or falls back to
    the saved snapshot for this camera if no file is provided.

    Returns PROPOSED ROIs — NOT persisted. The admin must accept proposals and
    save them via POST /api/roi/{camera_id} before they are stored.

    Query params:
      use_line_detection (bool): Snap candidate boxes to painted line markings
                                  via Canny + HoughLinesP (default: false).

    Honest constraints:
      Proposals reliably cover OCCUPIED spots. Empty spots are only detected
      when use_line_detection=True and markings are clearly visible. Always
      review proposals before accepting them.
    """
    if file is not None:
        allowed = (".jpg", ".jpeg", ".png", ".bmp")
        if not file.filename.lower().endswith(allowed):
            raise HTTPException(400, "Unsupported image format. Use JPG or PNG.")
        content = await file.read()
    else:
        snap_path = RoiStore.get_snapshot_path(camera_id)
        if snap_path is None:
            raise HTTPException(
                400,
                "No image uploaded and no snapshot found for this camera. "
                "Upload a reference image first.",
            )
        with open(snap_path, "rb") as fh:
            content = fh.read()

    nparr = np.frombuffer(content, np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(400, "Could not decode image")

    try:
        from src.inference.roi_proposer import propose_from_frames
        proposals = propose_from_frames(
            frames=[frame],
            camera_id=camera_id,
            use_line_detection=use_line_detection,
        )
    except Exception as exc:
        logger.error(f"ROI proposal failed for camera '{camera_id}': {exc}")
        raise HTTPException(500, f"Proposal failed: {exc}")

    return {
        "camera_id": camera_id,
        "proposals": proposals,
        "count": len(proposals),
        "warning": (
            "Proposals are based on vehicle detections (occupied spots). "
            "Empty spots may be missed. Review and edit all proposals before saving."
        ),
    }


# ── Camera registry endpoints ────────────────────────────

@app.get("/api/cameras", dependencies=[Depends(verify_api_key)])
def list_cameras():
    return camera_registry.get_all()

@app.post("/api/cameras", status_code=201, dependencies=[Depends(verify_api_key)])
async def add_camera(request: Request):
    body = await request.json()
    name = body.get("name", "").strip()
    source = body.get("source", "").strip()
    type_ = body.get("type", "usb")
    roi_camera_id = body.get("roi_camera_id", "").strip() or None

    if not name:
        raise HTTPException(400, "name is required")
    if not source:
        raise HTTPException(400, "source is required")
    if type_ not in ("usb", "rtsp", "file", "youtube"):
        raise HTTPException(400, "type must be usb, rtsp, file, or youtube")

    _validate_camera_source(source, type_)

    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-") or "cam"
    cam_id = f"{slug}-{uuid.uuid4().hex[:6]}"

    try:
        cam = camera_registry.add_camera(
            id=cam_id, name=name, source=source, type_=type_,
            roi_camera_id=roi_camera_id,
        )
    except ValueError as e:
        raise HTTPException(409, str(e))
    return cam

@app.patch("/api/cameras/{camera_id}", dependencies=[Depends(verify_api_key)])
async def update_camera(camera_id: str, request: Request):
    body = await request.json()
    name    = body.get("name",    "").strip() or None
    source  = body.get("source",  "").strip() or None
    type_   = body.get("type",    "").strip() or None
    roi_camera_id = body.get("roi_camera_id", "").strip() or None

    if type_ is not None and type_ not in ("usb", "rtsp", "file", "youtube"):
        raise HTTPException(400, "type must be usb, rtsp, file, or youtube")

    cam_current = camera_registry.get(camera_id)
    if cam_current is None:
        raise HTTPException(404, f"Camera '{camera_id}' not found")

    if source is not None:
        _validate_camera_source(source, type_ or cam_current["type"])

    cam = camera_registry.update_camera(
        camera_id, name=name, source=source, type_=type_, roi_camera_id=roi_camera_id
    )
    return cam

@app.delete("/api/cameras/{camera_id}", dependencies=[Depends(verify_api_key)])
def remove_camera(camera_id: str):
    if not camera_registry.remove_camera(camera_id):
        raise HTTPException(404, f"Camera '{camera_id}' not found")
    return {"deleted": camera_id}

@app.post("/api/cameras/{camera_id}/activate", dependencies=[Depends(verify_api_key)])
def activate_camera(camera_id: str):
    if camera_registry.get(camera_id) is None:
        raise HTTPException(404, f"Camera '{camera_id}' not found")
    ok = camera_registry.activate(camera_id, model_name=_resolve_model_name())
    if not ok:
        raise HTTPException(500, "Failed to activate camera")
    if _anomaly_enabled:
        p = camera_registry.get_processor(camera_id)
        if p:
            try:
                p.set_anomaly_detection(True)
            except Exception as e:
                logger.debug(f"Anomaly setup skipped for {camera_id}: {e}")
    return {"activated": camera_id}

@app.post("/api/cameras/{camera_id}/deactivate", dependencies=[Depends(verify_api_key)])
def deactivate_camera(camera_id: str):
    if camera_registry.get(camera_id) is None:
        raise HTTPException(404, f"Camera '{camera_id}' not found")
    camera_registry.deactivate(camera_id)
    return {"deactivated": camera_id}

# ═══════════════════════════════════════════════════════════════
# WebSocket — streams frames + metrics at ~20 FPS
# ═══════════════════════════════════════════════════════════════

def _ws_token_valid(token: str) -> bool:
    """Return True if the token is acceptable for WebSocket auth.
    When API_KEY is unset the check is skipped (open deployment).
    Set VITE_API_KEY in the frontend .env to pass the token automatically.
    """
    if not API_KEY:
        return True
    return hmac.compare_digest(token.encode(), API_KEY.encode())

@app.websocket("/ws/video")
async def video_ws(websocket: WebSocket, token: str = ""):
    if not _ws_token_valid(token):
        await websocket.close(code=4001)
        return
    await websocket.accept()
    proc = _get_processor()
    proc.start_processing()
    logger.info("WebSocket client connected")
    last_frame_seq = -1
    try:
        while True:
            proc = _get_processor()
            metrics = proc.get_metrics()
            payload = {"metrics": metrics}
            frame_b64, frame_seq = proc.get_frame_and_seq()
            if frame_b64 and frame_seq != last_frame_seq:
                last_frame_seq = frame_seq
                payload["frame"] = frame_b64
            await websocket.send_json(payload)
            await asyncio.sleep(0.05)
    except WebSocketDisconnect:
        logger.info("WebSocket client disconnected")
    except Exception as e:
        logger.error(f"WebSocket error: {e}")

@app.websocket("/ws/cameras/{camera_id}")
async def camera_ws(websocket: WebSocket, camera_id: str, token: str = ""):
    if not _ws_token_valid(token):
        await websocket.close(code=4001)
        return
    await websocket.accept()
    if camera_registry.get(camera_id) is None:
        await websocket.send_json({"error": "Camera not found"})
        await websocket.close()
        return
    if camera_registry.get_processor(camera_id) is None:
        await websocket.send_json({"type": "feed_unavailable", "reason": "Camera is not active"})
        await websocket.close()
        return
    logger.info(f"Camera WS connected: {camera_id}")
    no_frame_ticks = 0
    last_frame_seq = -1
    try:
        while True:
            proc = camera_registry.get_processor(camera_id)
            if proc is None:
                await websocket.send_json({"type": "feed_unavailable", "reason": "Camera feed stopped"})
                await websocket.close()
                break
            metrics = proc.get_metrics()
            payload = {"metrics": metrics}
            frame_b64, frame_seq = proc.get_frame_and_seq()
            if frame_b64 is None:
                no_frame_ticks += 1
                if no_frame_ticks >= 600:  # ~30 s with 0.05 s sleep
                    await websocket.send_json({"type": "feed_unavailable", "reason": "Video stream unavailable or timed out"})
                    await websocket.close()
                    break
            else:
                no_frame_ticks = 0
                if frame_seq != last_frame_seq:
                    last_frame_seq = frame_seq
                    payload["frame"] = frame_b64
            await websocket.send_json(payload)
            await asyncio.sleep(0.05)
    except WebSocketDisconnect:
        logger.info(f"Camera WS disconnected: {camera_id}")
    except Exception as e:
        logger.error(f"Camera WS error ({camera_id}): {e}")

# ── Edge → Hub ingest (hub side) ─────────────────────────
@app.post("/api/ingest/occupancy", dependencies=[Depends(verify_api_key)])
async def ingest_occupancy(request: Request):
    """Receive batched occupancy rows from an edge node and upsert into hub DB."""
    rows = await request.json()
    if not isinstance(rows, list):
        raise HTTPException(400, "Expected a JSON array of occupancy rows")
    inserted = db.upsert_occupancy_batch(rows)
    return {"inserted": inserted, "received": len(rows)}


@app.post("/api/ingest/alerts", dependencies=[Depends(verify_api_key)])
async def ingest_alerts(request: Request):
    """Receive batched alert rows from an edge node and upsert into hub DB."""
    rows = await request.json()
    if not isinstance(rows, list):
        raise HTTPException(400, "Expected a JSON array of alert rows")
    inserted = db.upsert_alerts_batch(rows)
    return {"inserted": inserted, "received": len(rows)}


# ── Data Augmentation Preview ─────────────────────────────
@app.post("/api/augment/preview", dependencies=[Depends(verify_api_key)])
async def augment_preview(request: Request):
    """Sample images from the dataset and return augmented versions as base64 JPEGs."""
    import random as _rnd

    body = await request.json()
    label    = body.get("label", "both")
    shadow_p = max(0.0, min(1.0, float(body.get("shadow_p", 0.5))))
    night    = bool(body.get("night", False))
    flip     = bool(body.get("flip", True))
    rotation = max(0, min(45, int(body.get("rotation", 15))))
    jitter   = max(0.0, min(1.0, float(body.get("jitter", 0.3))))
    count    = max(1, min(8, int(body.get("count", 6))))

    classes = ["occupied", "vacant"] if label == "both" else [label]
    candidates = []
    for cls in classes:
        cls_dir = config.DATA_DIR / cls
        if cls_dir.exists():
            for p in cls_dir.iterdir():
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".bmp"):
                    candidates.append((p, cls))

    if not candidates:
        raise HTTPException(400, "No dataset images found. Prepare the dataset first via the Controls panel.")

    chosen = _rnd.sample(candidates, min(count, len(candidates)))
    results = []

    for img_path, cls in chosen:
        try:
            img = Image.open(img_path).convert("RGB").resize((224, 224), Image.BILINEAR)
        except Exception:
            continue
        arr = np.array(img, dtype=np.float32)

        if flip and _rnd.random() < 0.5:
            arr = arr[:, ::-1, :].copy()

        if rotation > 0:
            angle = _rnd.uniform(-rotation, rotation)
            arr = np.array(
                Image.fromarray(arr.astype(np.uint8)).rotate(angle, expand=False),
                dtype=np.float32,
            )

        if jitter > 0:
            brightness = 1.0 + _rnd.uniform(-jitter, jitter)
            arr = np.clip(arr * brightness, 0, 255)
            contrast = 1.0 + _rnd.uniform(-jitter * 0.7, jitter * 0.7)
            mean = arr.mean()
            arr = np.clip((arr - mean) * contrast + mean, 0, 255)

        if shadow_p > 0 and _rnd.random() < shadow_p:
            _, w_s = arr.shape[:2]
            bw = _rnd.randint(w_s // 5, 3 * w_s // 5)
            x0 = _rnd.randint(0, w_s - bw)
            arr[:, x0:x0 + bw] *= _rnd.uniform(0.35, 0.65)

        if night:
            arr *= _rnd.uniform(0.15, 0.35)
            arr[:, :, 2] = np.clip(arr[:, :, 2] * 1.5, 0, 255)

        arr = np.clip(arr, 0, 255).astype(np.uint8)
        buf = io.BytesIO()
        Image.fromarray(arr).save(buf, format="JPEG", quality=85)
        import base64 as _b64
        results.append({"label": cls, "image": _b64.b64encode(buf.getvalue()).decode()})

    return {"images": results, "count": len(results)}


# ── Entry point ───────────────────────────────────────────
if __name__ == "__main__":
    uvicorn.run("main:app", host=config.HOST, port=config.PORT, reload=True)
